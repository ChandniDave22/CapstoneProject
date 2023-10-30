import openpyxl


def getRowCount(fileName, sheetName):
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.max_row

def readData(fileName, sheetName, rownum, colnum):
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum, column=colnum).value

def getData(datafilepath):
    row_count = getRowCount(datafilepath, "Sheet1")
    row_data_list = []
    for r in range(1, row_count+1):
      row_data = [readData(datafilepath, "Sheet1", r, 1), readData(datafilepath, "Sheet1", r, 2)]
      row_data_list.append(tuple(row_data))
    return row_data_list